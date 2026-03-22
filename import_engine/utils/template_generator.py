import io
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill
from openpyxl.comments import Comment


def generate_template(config) -> io.BytesIO:
    wb = Workbook()

    ws = wb.active
    ws.title = "Import Data"

    header_font = Font(bold=True, color="FFFFFF")
    required_fill = PatternFill(
        start_color="FF0000", end_color="FF0000", fill_type="solid"
    )
    optional_fill = PatternFill(
        start_color="FFC000", end_color="FFC000", fill_type="solid"
    )

    col_idx = 1

    for f_name, f_config in config.fields.items():
        if isinstance(f_config, list):
            rules = f_config
            label = f_name
        else:
            rules = f_config.get("rules", [])
            label = f_config.get("label", f_name)

        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font

        is_required = ("required" in rules) or (
            isinstance(f_config, dict) and f_config.get("required")
        )
        cell.fill = required_fill if is_required else optional_fill

        tips = [f"Field: {label}"]
        tips.append("Required" if is_required else "Optional")
        if isinstance(f_config, dict) and f_config.get("type") == "fk":
            tips.append(
                f"Foreign Key: lookups automatically resolved via {f_config.get('lookup', 'name')}."
            )

        cell.comment = Comment("\n".join(tips), "ImportEngine")

        # Support for Choices
        choices = []
        is_choice_or_fk = False

        if isinstance(f_config, dict):
            if f_config.get("type") == "choice":
                choices = f_config.get("choices", [])
                is_choice_or_fk = True
            elif f_config.get("fk"):
                from import_engine.domain.config_registry import (
                    get_config as get_registry_config,
                )

                fk_name = f_config.get("fk")
                fk_config = get_registry_config(fk_name)
                if fk_config:
                    lookup_field = getattr(
                        fk_config, "lookup_field", f_config.get("lookup", "name")
                    )
                    # Fetch first 100 choices for the dropdown to avoid huge Excel files
                    choices = list(
                        fk_config.model.objects.all().values_list(
                            lookup_field, flat=True
                        )[:100]
                    )
                    is_choice_or_fk = True

        if is_choice_or_fk and choices:
            # Create a custom named reference sheet for this specific field
            import re

            # Excel limits sheet names to 31 chars and prevents \ / * ? : [ ]
            # Replace invalid chars with space, then collapse multiple spaces, then trim
            safe_label = re.sub(r"[\\/*?:\[\]]", " ", label)
            safe_label = re.sub(r"\s+", " ", safe_label).strip()

            # Ensure " Ref" suffix fits within 31 chars
            base_name_limit = 31 - len(" Ref")
            base_name = safe_label[:base_name_limit].strip()

            safe_sheet_name = f"{base_name} Ref"

            # Handle duplicates if two fields sanitize to the same base name (e.g. "Rank/Role" and "Rank*Role")
            suffix_counter = 1
            while safe_sheet_name in wb.sheetnames:
                suffix = f" {suffix_counter}"
                new_limit = 31 - len(" Ref") - len(suffix)
                safe_sheet_name = f"{base_name[:new_limit].strip()} Ref{suffix}"
                suffix_counter += 1

            ref_ws = wb.create_sheet(title=safe_sheet_name)
            ref_ws.sheet_state = (
                "visible"  # visible for debugging, hide later if desired
            )

            ref_ws.cell(row=1, column=1, value=f"{label} Choices")
            choice_strs = []
            for r_idx, c_val in enumerate(choices, start=2):
                # Ensure atomic values and convert to string for Excel
                val = str(c_val[0] if isinstance(c_val, (list, tuple)) else c_val)
                choice_strs.append(val)
                ref_ws.cell(row=r_idx, column=1, value=val)

            # Use Named Ranges (Defined Names) for bulletproof Google Sheets compatibility
            from openpyxl.workbook.defined_name import DefinedName

            # If the CSV of choices is under 255 chars, we can embed it directly and skip the reference sheet formula entirely for better compatibility.
            csv_choices = ",".join(choice_strs)
            if len(csv_choices) < 255:
                # Wrap in double quotes for explicit literal list
                dv_formula = f'"{csv_choices}"'
                dv = DataValidation(
                    type="list",
                    formula1=dv_formula,
                    allow_blank=True,
                    showInputMessage=True,
                )
                ws.add_data_validation(dv)
            else:
                # Sheet names with spaces or special chars must be single-quoted in the range string
                safe_sheet_name_escaped = safe_sheet_name.replace("'", "''")
                range_str = f"'{safe_sheet_name_escaped}'!$A$2:$A${len(choices) + 1}"

                # Named ranges cannot contain spaces or special chars, only letters, numbers, and underscores
                safe_range_name = re.sub(r"[^A-Za-z0-9_]", "_", f_name)
                list_name = f"ChoicesList_{safe_range_name}_{col_idx}"

                defined_name = DefinedName(name=list_name, attr_text=range_str)
                if hasattr(wb.defined_names, "append"):
                    wb.defined_names.append(defined_name)
                else:
                    wb.defined_names.add(defined_name)

                # The formula must start with exactly one equal sign.
                dv_formula = f"={list_name}"

                dv = DataValidation(
                    type="list",
                    formula1=dv_formula,
                    allow_blank=True,
                    showInputMessage=True,
                )
                # Prevent Excel from wrapping formulas in literal quotes
                dv.quotePrefix = False
                ws.add_data_validation(dv)

            target_col_letter = ws.cell(row=1, column=col_idx).column_letter
            # Apply to the entire column below the header
            dv.add(f"{target_col_letter}2:{target_col_letter}1000")

            # Auto-adjust column width for the reference sheet
            ref_ws.column_dimensions["A"].width = max(len(label) + 10, 25)

        elif not is_choice_or_fk:
            dv = None
            min_val = None
            max_val = None

            for rule in rules:
                if rule.startswith("min:"):
                    min_val = rule.split(":")[1]
                elif rule.startswith("max:"):
                    max_val = rule.split(":")[1]

            if min_val is not None and max_val is not None:
                dv = DataValidation(
                    type="decimal",
                    operator="between",
                    formula1=min_val,
                    formula2=max_val,
                    allow_blank=not is_required,
                    showErrorMessage=True,
                    errorTitle="Invalid Range",
                    error=f"Value must be between {min_val} and {max_val}",
                )
            elif min_val is not None:
                dv = DataValidation(
                    type="decimal",
                    operator="greaterThanOrEqual",
                    formula1=min_val,
                    allow_blank=not is_required,
                    showErrorMessage=True,
                    errorTitle="Minimum Value",
                    error=f"Value must be at least {min_val}",
                )
            elif max_val is not None:
                dv = DataValidation(
                    type="decimal",
                    operator="lessThanOrEqual",
                    formula1=max_val,
                    allow_blank=not is_required,
                    showErrorMessage=True,
                    errorTitle="Maximum Value",
                    error=f"Value must be at most {max_val}",
                )
            elif isinstance(f_config, dict) and f_config.get("type") == "date":
                dv = DataValidation(
                    type="date",
                    operator="greaterThan",
                    formula1='"1900-01-01"',
                    allow_blank=not is_required,
                    showErrorMessage=True,
                    errorTitle="Invalid Date",
                    error="Please enter a valid date",
                )

            if dv:
                ws.add_data_validation(dv)
                target_col_letter = ws.cell(row=1, column=col_idx).column_letter
                dv.add(f"{target_col_letter}2:{target_col_letter}1000")

        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(
            len(label) + 5, 15
        )
        col_idx += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
