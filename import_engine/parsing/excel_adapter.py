from openpyxl import load_workbook
from import_engine.parsing.base_adapter import BaseParserAdapter

class ExcelAdapter(BaseParserAdapter):
    def __init__(self, file_path_or_buffer):
        super().__init__(file_path_or_buffer)
        # Using read_only=True significantly reduces memory footprint for large files
        self.workbook = load_workbook(filename=self.source, read_only=True, data_only=True)
        
        # Explicitly read from the primary data sheet rather than whatever sheet the user happened to leave active
        if "Import Data" in self.workbook.sheetnames:
            self.sheet = self.workbook["Import Data"]
        else:
            self.sheet = self.workbook.active
            
        self._headers = None

    def get_headers(self) -> list[str]:
        if not self._headers:
            for row in self.sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                self._headers = [str(cell) if cell else "" for cell in row]
                break
        return self._headers or []

    def iter_rows(self, start_row: int = 1, end_row: int = None):
        headers = self.get_headers()
        # +1 because Excel rows are 1-indexed and header is row 1
        # Data starts at row 2, which corresponds to start_row=1 in terms of data index
        excel_start = start_row + 1
        excel_end = (end_row + 1) if end_row else None
        
        for excel_idx, row in enumerate(self.sheet.iter_rows(min_row=excel_start, max_row=excel_end, values_only=True), start=start_row):
            row_dict = dict(zip(headers, row))
            yield excel_idx, row_dict

    def chunked_read(self, chunk_size=1000):
        chunk = []
        # data iteration starts at 1
        for idx, row_dict in self.iter_rows():
            chunk.append((idx, row_dict))
            if len(chunk) == chunk_size:
                yield chunk
                chunk = []
        if chunk:
            yield chunk

    def close(self):
        self.workbook.close()
